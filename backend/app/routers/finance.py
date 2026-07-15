import io
from collections import defaultdict
from datetime import date
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.dependencies import require_admin
from app.database import get_db
from app.models.owner_settlement import OwnerSettlement
from app.models.reservation import Reservation, ReservationStatus
from app.models.vehicle import Vehicle

router = APIRouter(prefix="/api/finance", tags=["finance"], redirect_slashes=False)

ACTIVE_STATUSES = {
    ReservationStatus.lead,
    ReservationStatus.quoted,
    ReservationStatus.deposit_received,
    ReservationStatus.reserved,
    ReservationStatus.confirmed,
}


@router.get("/summary", dependencies=[Depends(require_admin)])
def finance_summary(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    today = date.today()
    current_year = today.year

    # Year-level revenue (always full years, ignores date range filter)
    def _year_revenue(year: int) -> float:
        result = db.query(func.coalesce(func.sum(Reservation.total_amount), 0)).filter(
            Reservation.status == ReservationStatus.completed,
            func.extract("year", Reservation.event_date) == year,
        ).scalar()
        return float(result or 0)

    revenue_this_year = _year_revenue(current_year)
    revenue_last_year = _year_revenue(current_year - 1)
    yoy_change_pct = (
        round((revenue_this_year - revenue_last_year) / revenue_last_year * 100, 1)
        if revenue_last_year > 0 else None
    )

    # Period-filtered metrics
    eff_from = date_from or today.replace(day=1)
    eff_to = date_to or today

    period_filters = [
        Reservation.event_date >= eff_from,
        Reservation.event_date <= eff_to,
    ]

    completed_period = (
        db.query(Reservation, Vehicle)
        .outerjoin(Vehicle, Reservation.vehicle_id == Vehicle.id)
        .filter(Reservation.status == ReservationStatus.completed, *period_filters)
        .all()
    )
    revenue_period = float(sum(r.total_amount for r, _ in completed_period) or 0)
    company_revenue_period = float(sum(
        r.total_amount if (v and v.is_company_owned) else r.total_amount * Decimal("0.30")
        for r, v in completed_period
    ) or 0)

    deposits_received_period = float(
        db.query(func.coalesce(func.sum(Reservation.deposit_paid), 0))
        .filter(
            Reservation.status != ReservationStatus.cancelled,
            *period_filters,
        )
        .scalar() or 0
    )

    # Outstanding balance (all-time, active reservations only)
    active_res = db.query(Reservation).filter(
        Reservation.status.in_(ACTIVE_STATUSES)
    ).all()
    outstanding_balance_total = float(
        sum(r.remaining_balance for r in active_res) or 0
    )

    # Pending owner payments (unsettled OwnerSettlements)
    pending_owner_payments = float(
        db.query(func.coalesce(func.sum(OwnerSettlement.owner_amount), 0))
        .filter(OwnerSettlement.status == "pending")
        .scalar() or 0
    )

    return {
        "revenue_this_year": revenue_this_year,
        "revenue_last_year": revenue_last_year,
        "yoy_change_pct": yoy_change_pct,
        "revenue_period": revenue_period,
        "company_revenue_period": round(company_revenue_period),
        "deposits_received_period": deposits_received_period,
        "outstanding_balance_total": outstanding_balance_total,
        "pending_owner_payments": pending_owner_payments,
    }


@router.get("/charts/owner-revenue", dependencies=[Depends(require_admin)])
def owner_revenue(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    # Join reservations with vehicles to get owner_name
    query = (
        db.query(Reservation, Vehicle)
        .outerjoin(Vehicle, Reservation.vehicle_id == Vehicle.id)
        .filter(Reservation.status == ReservationStatus.completed)
    )
    if date_from:
        query = query.filter(Reservation.event_date >= date_from)
    if date_to:
        query = query.filter(Reservation.event_date <= date_to)

    rows = query.all()

    owner_map: dict = {}
    for reservation, vehicle in rows:
        is_company = vehicle.is_company_owned if vehicle else False
        owner_name = (
            "Camino a mi Boda" if is_company
            else (vehicle.owner_name if vehicle and vehicle.owner_name else "Sin propietario")
        )
        if owner_name not in owner_map:
            owner_map[owner_name] = {"owner_name": owner_name, "completed_count": 0, "total_revenue": 0.0, "is_company": is_company}
        owner_map[owner_name]["completed_count"] += 1
        owner_map[owner_name]["total_revenue"] += float(reservation.total_amount)

    owners = []
    for entry in owner_map.values():
        rev = entry["total_revenue"]
        is_co = entry["is_company"]
        owners.append({
            "owner_name": entry["owner_name"],
            "completed_count": entry["completed_count"],
            "total_revenue": rev,
            "owner_amount": 0 if is_co else round(rev * 0.70),
            "company_amount": round(rev) if is_co else round(rev * 0.30),
        })

    owners.sort(key=lambda x: x["total_revenue"], reverse=True)
    return {"owners": owners}


@router.get("/charts/deposits", dependencies=[Depends(require_admin)])
def deposits_chart(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    today = date.today()

    # Default: last 12 months
    if not date_from and not date_to:
        twelve_ago = today.replace(day=1)
        month = twelve_ago.month - 11
        year = twelve_ago.year
        if month <= 0:
            month += 12
            year -= 1
        eff_from = date(year, month, 1)
        eff_to = today
    else:
        eff_from = date_from
        eff_to = date_to

    filters = [Reservation.status != ReservationStatus.cancelled]
    if eff_from:
        filters.append(Reservation.event_date >= eff_from)
    if eff_to:
        filters.append(Reservation.event_date <= eff_to)

    rows = db.query(Reservation).filter(*filters).all()

    monthly: dict = {}
    for r in rows:
        key = r.event_date.strftime("%Y-%m")
        if key not in monthly:
            monthly[key] = {"month": key, "deposits_received": 0.0, "remaining_balance": 0.0}
        monthly[key]["deposits_received"] += float(r.deposit_paid)
        monthly[key]["remaining_balance"] += float(r.remaining_balance)

    data = sorted(
        [{"month": k, "deposits_received": round(v["deposits_received"]), "remaining_balance": round(v["remaining_balance"])}
         for k, v in monthly.items()],
        key=lambda x: x["month"],
    )
    return {"data": data}


@router.get("/aging", dependencies=[Depends(require_admin)])
def aging(db: Session = Depends(get_db)):
    today = date.today()

    rows = (
        db.query(Reservation)
        .filter(
            Reservation.status.notin_([ReservationStatus.cancelled, ReservationStatus.completed]),
        )
        .order_by(Reservation.event_date.asc())
        .all()
    )

    items = []
    for r in rows:
        rb = float(r.remaining_balance)
        if rb <= 0:
            continue
        days_to_event = (r.event_date - today).days
        items.append({
            "id": r.id,
            "reservation_number": r.reservation_number,
            "display_customer": r.display_customer,
            "event_date": r.event_date.isoformat(),
            "days_to_event": days_to_event,
            "total_amount": float(r.total_amount),
            "deposit_paid": float(r.deposit_paid),
            "remaining_balance": rb,
            "status": r.status if isinstance(r.status, str) else r.status.value,
        })

    return {"items": items}


def _months_set(start: date, end: date) -> set:
    months: set = set()
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.add(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def _vehicle_perf(db: Session, eff_from: date, eff_to: date) -> tuple[list, int]:
    all_months = _months_set(eff_from, eff_to)
    total_months = len(all_months)

    vehicles = db.query(Vehicle).order_by(Vehicle.display_order).all()

    completed = (
        db.query(Reservation)
        .filter(
            Reservation.status == ReservationStatus.completed,
            Reservation.event_date >= eff_from,
            Reservation.event_date <= eff_to,
        )
        .all()
    )

    res_by_vehicle: dict = defaultdict(list)
    for r in completed:
        if r.vehicle_id:
            res_by_vehicle[r.vehicle_id].append(r)

    rows = []
    for v in vehicles:
        rsvs = res_by_vehicle.get(v.id, [])
        total_rev = float(sum(r.total_amount for r in rsvs) or 0)
        count = len(rsvs)
        active_months = {r.event_date.strftime("%Y-%m") for r in rsvs}
        idle = max(0, total_months - len(active_months))
        company_share = total_rev if v.is_company_owned else total_rev * 0.30

        name_parts = [v.brand]
        if v.model_line:
            name_parts.append(v.model_line)
        if v.year:
            name_parts.append(str(v.year))
        name = " ".join(name_parts)
        if v.color:
            name += f" — {v.color}"

        rows.append({
            "vehicle_id": v.id,
            "name": name,
            "owner": "Camino a mi Boda" if v.is_company_owned else (v.owner_name or "—"),
            "is_company_owned": v.is_company_owned,
            "completed_events": count,
            "total_revenue": round(total_rev),
            "company_share": round(company_share),
            "avg_ticket": round(total_rev / count) if count else 0,
            "idle_months": idle,
        })

    rows.sort(key=lambda x: x["total_revenue"], reverse=True)
    return rows, total_months


@router.get("/charts/vehicle-revenue", dependencies=[Depends(require_admin)])
def vehicle_revenue_chart(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    today = date.today()
    if not date_from and not date_to:
        m, y = today.month - 11, today.year
        if m <= 0:
            m += 12
            y -= 1
        eff_from, eff_to = date(y, m, 1), today
    else:
        eff_from = date_from or date(today.year - 1, today.month, 1)
        eff_to = date_to or today

    rows, total_months = _vehicle_perf(db, eff_from, eff_to)
    return {"vehicles": rows, "total_months": total_months}


@router.get("/export", dependencies=[Depends(require_admin)])
def export_excel(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook  # lazy import — only needed on export

    today = date.today()

    res_query = db.query(Reservation).order_by(Reservation.event_date.desc())
    if date_from:
        res_query = res_query.filter(Reservation.event_date >= date_from)
    if date_to:
        res_query = res_query.filter(Reservation.event_date <= date_to)
    reservations = res_query.all()

    if not date_from and not date_to:
        m, y = today.month - 11, today.year
        if m <= 0:
            m += 12
            y -= 1
        perf_from, perf_to = date(y, m, 1), today
    else:
        perf_from = date_from or date(today.year - 1, today.month, 1)
        perf_to = date_to or today

    vehicle_rows, total_months = _vehicle_perf(db, perf_from, perf_to)

    settlements = db.query(OwnerSettlement).order_by(OwnerSettlement.created_at.desc()).all()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Reservaciones"
    ws1.append(["#", "Número", "Cliente", "Fecha evento", "Vehículo", "Conductor", "Estado",
                 "Total (COP)", "Abono (COP)", "Saldo (COP)"])
    for i, r in enumerate(reservations, 1):
        ws1.append([
            i,
            r.reservation_number,
            r.display_customer,
            r.event_date.isoformat(),
            r.display_vehicle,
            r.display_driver,
            r.status if isinstance(r.status, str) else r.status.value,
            float(r.total_amount),
            float(r.deposit_paid),
            float(r.remaining_balance),
        ])

    ws2 = wb.create_sheet("Rendimiento por vehículo")
    ws2.append(["Vehículo", "Propietario", "Eventos completados", "Ingresos (COP)",
                 "Parte empresa (COP)", "Ticket promedio (COP)", "Meses inactivos",
                 f"de {total_months} meses en rango"])
    for row in vehicle_rows:
        ws2.append([
            row["name"],
            row["owner"],
            row["completed_events"],
            row["total_revenue"],
            row["company_share"],
            row["avg_ticket"],
            row["idle_months"],
            total_months,
        ])

    ws3 = wb.create_sheet("Liquidaciones")
    ws3.append(["Número", "Fecha", "Vehículo", "Propietario", "Reservación",
                 "Valor evento (COP)", "Parte propietario (COP)", "% propietario",
                 "Parte empresa (COP)", "Estado", "Notas"])
    for s in settlements:
        v_name = "—"
        if s.vehicle:
            parts = [s.vehicle.brand]
            if s.vehicle.model_line:
                parts.append(s.vehicle.model_line)
            v_name = " ".join(parts)
        ws3.append([
            s.settlement_number,
            s.created_at.date().isoformat() if s.created_at else "",
            v_name,
            s.owner.full_name if s.owner else "—",
            s.reservation.reservation_number if s.reservation else "—",
            float(s.reservation_value),
            float(s.owner_amount),
            s.owner_percentage,
            float(s.company_amount),
            s.status,
            s.notes or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="camino-export.xlsx"'},
    )
